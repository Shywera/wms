"""Čitanje zahtjevnice iz datoteke → popis redaka {sifra, naziv, jedinica, kolicina, rn}.

Podržano: .xlsx (openpyxl), .csv (stdlib), .pdf (pdfplumber, best-effort).
Brojevi su u hrvatskom formatu: točka = tisućice, zarez = decimale
(npr. "159.000" = 159000 araka, "2,5" = 2.5 kg).
Nakon uvoza skladištar UVIJEK pregleda stavke prije izdavanja (review), pa manje
greške parsiranja (osobito PDF) nisu opasne.
"""
from __future__ import annotations

import csv
import io
import re


def parse_hr_broj(s) -> float:
    """'159.000' -> 159000.0 ; '2,5' -> 2.5 ; '46' -> 46.0 ; '' -> 0.0"""
    if s is None:
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    t = str(s).strip().replace("\xa0", "").replace(" ", "")
    if not t:
        return 0.0
    if "," in t:                      # zarez = decimalni → makni tisućice (.), zarez u točku
        t = t.replace(".", "").replace(",", ".")
    else:                             # samo točke = tisućice (cijeli broj araka)
        t = t.replace(".", "")
    try:
        return float(t)
    except ValueError:
        m = re.search(r"-?\d+(\.\d+)?", t)
        return float(m.group()) if m else 0.0


def _cisti_sifru(v) -> str:
    return re.sub(r"\D", "", str(v or ""))     # samo znamenke (šifra zna biti razlomljena)


# ── Zaglavlje (mapiranje stupaca po nazivu) ──────────────────────────────────
def _idx(headers: list[str], *kljucevi: str) -> int | None:
    low = [(h or "").strip().lower() for h in headers]
    for k in kljucevi:
        for i, h in enumerate(low):
            if k in h:
                return i
    return None


def _mapa_stupaca(headers: list[str]) -> dict | None:
    i_sifra = _idx(headers, "šifra", "sifra")
    i_naziv = _idx(headers, "naziv")
    i_jed = _idx(headers, "mjerna", "jedinica")
    i_kol = _idx(headers, "potrebna", "količina", "kolicina")
    i_rn = _idx(headers, "rn", "radni")
    if i_sifra is None or i_kol is None:
        return None
    return {"sifra": i_sifra, "naziv": i_naziv, "jedinica": i_jed, "kolicina": i_kol, "rn": i_rn}


def _red_iz_celija(cells: list, m: dict) -> dict | None:
    def g(key):
        i = m.get(key)
        return cells[i] if (i is not None and i < len(cells)) else None
    sifra = _cisti_sifru(g("sifra"))
    if not sifra:
        return None
    kol = parse_hr_broj(g("kolicina"))
    jed = str(g("jedinica") or "").strip().lower()
    naziv = str(g("naziv") or "").strip() or None
    rn = str(g("rn") or "").strip() or None
    return {"sifra": sifra, "naziv": naziv, "jedinica": jed, "kolicina": kol, "rn": rn}


# ── XLSX ─────────────────────────────────────────────────────────────────────
def _parse_xlsx(data: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    return _iz_redaka(rows)


# ── CSV ──────────────────────────────────────────────────────────────────────
def _parse_csv(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:2000]
    delim = ";" if sample.count(";") >= sample.count(",") else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    return _iz_redaka(rows)


def _iz_redaka(rows: list[list]) -> list[dict]:
    """Nađi zaglavlje pa pročitaj podatkovne retke (za xlsx/csv)."""
    m = None
    out = []
    for r in rows:
        cells = list(r)
        if m is None:
            cand = _mapa_stupaca([str(c) if c is not None else "" for c in cells])
            if cand:
                m = cand
            continue
        red = _red_iz_celija(cells, m)
        if red:
            out.append(red)
    return out


# ── PDF (best-effort) ────────────────────────────────────────────────────────
def _parse_pdf(data: bytes) -> list[dict]:
    import pdfplumber
    out: list[dict] = []
    m = None
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for r in table:
                    cells = [(c or "").replace("\n", " ").strip() for c in r]
                    if m is None:
                        low = " ".join(cells).lower()
                        if ("šifra" in low or "sifra" in low) and ("naziv" in low):
                            m = _mapa_stupaca(cells)
                        continue
                    red = _red_iz_celija(cells, m)
                    if red:
                        out.append(red)
    return out


# ── Javni ulaz ───────────────────────────────────────────────────────────────
def parsiraj(filename: str, data: bytes) -> tuple[list[dict], str | None]:
    """Vrati (redci, greska|None). Redak = {sifra, naziv, jedinica, kolicina, rn}."""
    ime = (filename or "").lower()
    try:
        if ime.endswith(".xlsx"):
            redci = _parse_xlsx(data)
        elif ime.endswith(".csv"):
            redci = _parse_csv(data)
        elif ime.endswith(".pdf"):
            redci = _parse_pdf(data)
        else:
            return [], "Nepodržan format. Uvezi .xlsx, .csv ili .pdf."
    except Exception as e:  # noqa: BLE001
        return [], f"Greška pri čitanju datoteke: {e}"

    if not redci:
        if ime.endswith(".pdf"):
            return [], ("PDF nije pouzdano pročitan (tablica nije prepoznata). "
                        "Preporuka: izvezi zahtjevnicu iz Pauka kao Excel (.xlsx) pa uvezi nju.")
        return [], "Nije pronađena nijedna stavka (provjeri da datoteka ima stupce Šifra i Potrebna količina)."
    return redci, None
